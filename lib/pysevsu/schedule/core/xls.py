import asyncio
import openpyxl

from typing import Any
from typing import List
from typing import Dict
from typing import Any
from typing import Optional
from io import BytesIO

from ..utilites.logger import log


class ExcelFile:
    def __init__(self, file: BytesIO):
        self.file = openpyxl.load_workbook(
            filename=file, 
            read_only=True
        ) 

    @property
    def sheetnames(self) -> List[str]:
        return self.file.sheetnames

    async def run_worksheets_stream(self) -> object:
        for sheetname in self.sheetnames:
            if sheetname.startswith("уч.н."):
                try:
                    yield Worksheet(self.file[sheetname], sheetname)
                except RuntimeError:
                    ...


class Worksheet:
    def __init__(
        self, 
        content: ExcelFile, 
        title: Optional[str] = ...
    ):
        self.content = content
        self.title = title
        self.data = self._load_cache()

        self._result: Dict[str, Any] = dict()
        self._tmp: Dict[str, Any] = dict()
        
        if self._max_row < 50:
            raise RuntimeError("Invalid size of the worksheet.")

    def _load_cache(self) -> List[List[Any]]:
        return [
            [cell.value for cell in row] 
            for row in self.content.rows
        ] # TODO: Сделать динамическую выгрузку с последющим подхватом

    @property
    def _max_row(self) -> int:
        return len(self.data)
    
    @property
    def _max_col(self) -> int:
        return len(self.data[0])

    def _cell(self, row: int, column: int) -> Optional[str]:
        try: 
            return self.data[row][column]
        except IndexError: 
            return None
        
    def _get_column_title(self, col: int):
        title = self._cell(4, col)
        if (
            not title or 
            title.startswith("подгруппа")
        ):
            title = self._cell(5, col)
        return title
    
    def _process_column_groups(self, col):
        cell = self._cell(3, col)
        if cell:
            self._result["group"] = cell

    @staticmethod
    def _value_validation(title: str, value: Any):
        return True if value and title != value else False
    
    def _process_lesson_information(self, title: str, value: Any):
        if title == "День":
            self._result["weekday"] = value
        if title == "Дата":
            self._result["date"] = value
        if title == "№занятия":
            self._result["number"] = value
        if title == "Время":
            self._result["start_time"] = value
    
    def _process_lesson_data(self, title: str, value: Any):
        if title == "Занятие":
            self._tmp["lessons"] = value.splitlines()
            self._tmp.pop("types", None)
            self._tmp.pop("classrooms", None)
        if title == "Тип":
            self._tmp["types"] = value.splitlines()
        if title == "Аудитория":
            self._tmp["classrooms"] = value.splitlines()

    async def _run_cell_processing(self):
        len_: int = len(self._tmp.get("lessons"))
        for index in range(len_):
            title, teacher = self._parse_lesson_line(
                self._tmp["lessons"][index]
            )

            self._result["title"] = title
            self._result["teacher"] = teacher

            if len_ == len(self._tmp["classrooms"]):
                self._result["classroom"] = self._tmp["classrooms"][index]
            else: 
                self._result["classroom"] = ''.join(self._tmp["classrooms"])

            if len_ == len(self._tmp["types"]):
                self._result["type"] = self._tmp["types"][index]
            else:
                self._result["type"] = ''.join(self._tmp["types"])
            
            yield self._result

    async def run_data_stream(self) -> object:
        for row in range(self._max_row):
            for col in range(self._max_col):
                self._process_column_groups(col)
                title = self._get_column_title(col)
                value = self._cell(row, col)
        
                if not self._value_validation(title, value):
                    continue
                
                self._process_lesson_information(title, value)
                self._process_lesson_data(title, value)
                
                if title == "Аудитория":
                    async for record in self._run_cell_processing():
                        yield record
                        self._result.clear()

    @staticmethod
    def _parse_lesson_line(str_: str):
        if ', ' in str_:
            tmp = str_.split(', ')
            title = ' '.join(tmp[:-1])
            teacher = tmp[-1]
        else:
            title = str_
            teacher = ""
        return title, teacher

    def get_dates_of_the_week(self) -> Dict[str, str]:
        return {
            "start_date" : self._cell(6, 1), 
            "end_date" : self._cell(46, 1)
        }


async def test():
    from .web import async_xls_request

    url = "https://www.sevsu.ru/univers/shedule/download.php?file=IZWT%2BjTApz5TmxgylFGt4A%3D%3D"
    xls_content = await async_xls_request(url)
    xls = ExcelFile(xls_content)

    async for sheet in xls.run_worksheets_stream():
        async for i in sheet.run_data_stream():
            print(i)


if __name__ == "__main__":
    asyncio.run(test())