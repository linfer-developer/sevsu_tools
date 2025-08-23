import asyncio
import openpyxl
from typing import Any
from typing import List
from core.data_manager import WorksheetCacheHandler


class ExcelFile:

    """The ExcelFile class provides tools for working with Excel a file from 
    the SevSU schedule website.
    It is used for processing Excel files with the university schedule.

    Attributes:
        file: IOBytes
            IOBytes object (Excel file).
        website_path: str
            The path to the Excel file on the website with the SevGU schedule.

    Methods:
        sheetnames()
            Returns a list of Excel file worksheets.
        async_import()
            Asynchronous import to the database. Creates Worksheet class 
            objects (a worker object an Excel sheet of a file) and collects 
            data from it.
        import_()
            Synchronous import to the database. Creates Worksheet class 
            objects (a worker object an Excel sheet of a file) and collects 
            data from it.
    """
    def __init__(self, file: Any, website_path: str):
        self.file = file # IOBytes object only
        self.website_path = website_path 
        self.excel = openpyxl.load_workbook(self.file, read_only=True)

    @property
    def sheetnames(self) -> List[str]:
        """Returns a list of worksheet names."""
        return self.excel.sheetnames

    async def async_import(self) -> None:
        """Asynchronous import to the database. Creates Worksheet objects and 
        collects data from them.

        It is commonly used in other asynchronous functions and methods. 
        """
        tasks = list()
        for sheetname in self.sheetnames:
            if sheetname.startswith("уч.н."):
                tasks.append(Worksheet(
                    excel=self, worksheet_name=sheetname
                ).import_())
        await asyncio.gather(*tasks)


class Worksheet:

    """The class provides tools for working with an Excel file worksheet.

    Used for caching data from the Excel worksheet and their subsequent data 
    processing of the data_manager module.
    
    Attributes:
        excel: ExcelFile
            The created Excel file object (ExcelFile).
        worksheet_name: str
            The name of the worksheet that should be used work must be done.

    Methods:
        async_import()
            Asynchronous collection of cached data using the data_manager 
            module.
        import_()
            Synchronous collection of cached data using the data_manager 
            module.
    """
    def __init__(
        self, 
        excel: ExcelFile, 
        worksheet_name: str
    ):
        self.excel = excel 
        self.worksheet_name = worksheet_name
        self.sheet = self.excel.excel[self.worksheet_name]
        self.worksheet_cache = self._load_cache()

    def _load_cache(self) -> List[List[str]]:
        """Caching data from an Excel file sheet."""
        return [[cell.value for cell in row] for row in self.sheet.rows]

    async def import_(self) -> None:
        """The method essentially creates an object of the 
        WorksheetCacheHeandler class, which in turn imports the entire 
        schedule. 

        Additional data is transmitted to it, which may be required or 
        desirable for import.
        """
        additional_data = self.excel.website_path.split("/")
        worksheet_number = int(self.worksheet_name.split()[-1])
        handler = WorksheetCacheHandler(
            data=self.worksheet_cache,
            additional_data={
                "week_number" : worksheet_number, 
                "study_form" : additional_data[0], 
                "institute" : additional_data[1], 
                "semester" : additional_data[2], 
                "full_course_name" : additional_data[3] 
            }
        )
        await handler.async_import_data()